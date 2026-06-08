import openpyxl

wb = openpyxl.load_workbook("Assessment of components operating in the creep range_MFC.xlsx", data_only=False)

with open("excel_formulas.txt", "w", encoding="utf-8") as f:
    for sheet_name in wb.sheetnames:
        f.write(f"\n==========================================\n")
        f.write(f"SHEET: {sheet_name}\n")
        f.write(f"==========================================\n")
        ws = wb[sheet_name]
        for r_idx in range(1, ws.max_row + 1):
            row_vals = []
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                val = cell.value
                if val is not None:
                    if isinstance(val, str) and val.startswith("="):
                        row_vals.append(f"Col {c_idx} [Formula: {val}]")
                    else:
                        row_vals.append(f"Col {c_idx} [Val: {val}]")
            if row_vals:
                f.write(f"Row {r_idx:2d}: " + " | ".join(row_vals) + "\n")
print("Done writing to excel_formulas.txt")
